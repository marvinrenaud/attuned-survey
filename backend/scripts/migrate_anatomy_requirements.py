#!/usr/bin/env python3
"""
Migrate anatomy requirements from spreadsheet to database.

Reads columns 8-9 (activePlayer Must Have, partnerPlayer Must Have) from
Consolidated_ToD_Activities (20).xlsx and updates activities.required_bodyparts
"""
import sys
import os
import argparse
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from src.main import app
from src.extensions import db
from src.models.activity import Activity

# Column indices (1-indexed in Excel)
DESCRIPTION_COL = 2
ACTIVE_ANATOMY_COL = 8
PARTNER_ANATOMY_COL = 9


def parse_anatomy(value):
    """Parse anatomy cell value to list of valid body parts."""
    if not value or str(value).strip().lower() in ['null', 'none', '', 'n/a']:
        return []
    
    # Split by comma if multiple
    parts = [p.strip().lower() for p in str(value).split(',')]
    valid = ['penis', 'vagina', 'breasts']
    return [p for p in parts if p in valid]


def normalize_description(desc):
    """Normalize description for matching (lowercase, strip whitespace)."""
    if not desc:
        return ''
    return str(desc).strip().lower()


def main():
    parser = argparse.ArgumentParser(description='Migrate anatomy requirements from spreadsheet')
    parser.add_argument('--execute', action='store_true', help='Execute migration (default is dry-run)')
    parser.add_argument('--spreadsheet', type=str, default='../../Consolidated_ToD_Activities (20).xlsx',
                       help='Path to spreadsheet file')
    args = parser.parse_args()
    
    dry_run = not args.execute
    
    if dry_run:
        print("=" * 70)
        print("DRY RUN MODE - No changes will be made")
        print("=" * 70)
    else:
        print("=" * 70)
        print("EXECUTE MODE - Changes will be applied")
        print("=" * 70)
    
    spreadsheet_path = Path(__file__).parent / args.spreadsheet
    
    if not spreadsheet_path.exists():
        print(f"❌ Spreadsheet not found: {spreadsheet_path}")
        sys.exit(1)
    
    print(f"\n📂 Reading spreadsheet: {spreadsheet_path.name}")
    
    # Load spreadsheet
    wb = openpyxl.load_workbook(str(spreadsheet_path))
    ws = wb['Consolidated Activities']
    
    print(f"   Rows: {ws.max_row}")
    print(f"   Columns: {ws.max_column}")
    
    with app.app_context():
        # Build lookup map: description → anatomy requirements
        anatomy_map = {}
        
        # Read from spreadsheet (skip header row)
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]
            
            # Get description (Column 2)
            description = row[DESCRIPTION_COL - 1].value
            if not description:
                continue
            
            # Get anatomy requirements
            active_anatomy = parse_anatomy(row[ACTIVE_ANATOMY_COL - 1].value)
            partner_anatomy = parse_anatomy(row[PARTNER_ANATOMY_COL - 1].value)
            
            # Skip if no requirements
            if not active_anatomy and not partner_anatomy:
                continue
            
            # Store normalized
            desc_key = normalize_description(description)
            anatomy_map[desc_key] = {
                "active": active_anatomy,
                "partner": partner_anatomy
            }
        
        print(f"\n📊 Found {len(anatomy_map)} activities with anatomy requirements")
        
        # Update database activities
        updated_count = 0
        not_found_count = 0
        already_set_count = 0
        
        activities = Activity.query.filter_by(is_active=True).all()
        print(f"   Checking {len(activities)} active activities in database...")
        
        for activity in activities:
            # Get activity description from script
            steps = activity.script.get('steps', [])
            if not steps:
                continue
            
            activity_desc = steps[0].get('do', '')
            desc_key = normalize_description(activity_desc)
            
            # Look up anatomy requirements
            anatomy_req = anatomy_map.get(desc_key)
            
            if anatomy_req:
                # Check if already set correctly
                current = activity.required_bodyparts or {"active": [], "partner": []}
                if (current.get('active') == anatomy_req['active'] and 
                    current.get('partner') == anatomy_req['partner']):
                    already_set_count += 1
                    continue
                
                if dry_run:
                    print(f"\n   Would update Activity {activity.activity_id}:")
                    print(f"     Description: {activity_desc[:60]}...")
                    print(f"     Current: {current}")
                    print(f"     New: {anatomy_req}")
                else:
                    activity.required_bodyparts = anatomy_req
                
                updated_count += 1
            else:
                # No anatomy requirements for this activity (that's ok)
                pass
        
        if not dry_run and updated_count > 0:
            db.session.commit()
            print(f"\n✅ Database updated successfully")
        
        print("\n" + "=" * 70)
        print("MIGRATION SUMMARY")
        print("=" * 70)
        print(f"   Activities scanned: {len(activities)}")
        print(f"   {'Would update' if dry_run else 'Updated'}: {updated_count}")
        print(f"   Already correct: {already_set_count}")
        print(f"   No requirements: {len(activities) - updated_count - already_set_count}")
        
        if dry_run:
            print(f"\nTo apply changes, run with --execute flag")
        
        print("=" * 70)


if __name__ == '__main__':
    main()

