"""Import activities from XLSX/CSV spreadsheet to database with enhanced metadata support."""
import sys
import os
import csv
import json
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.extensions import db
from src.models.activity import Activity, ALLOWED_BOUNDARIES, ALLOWED_BODYPARTS
from src.main import app


# Mapping from intimacy levels to intensity and rating
INTIMACY_LEVEL_MAP = {
    'L1': {'intensity': 1, 'rating': 'G'},
    'L2': {'intensity': 1, 'rating': 'G'},
    'L3': {'intensity': 1, 'rating': 'G'},
    'L4': {'intensity': 2, 'rating': 'R'},
    'L5': {'intensity': 3, 'rating': 'R'},
    'L6': {'intensity': 3, 'rating': 'R'},
    'L7': {'intensity': 4, 'rating': 'R'},
    'L8': {'intensity': 4, 'rating': 'X'},
    'L9': {'intensity': 5, 'rating': 'X'},
}


def generate_activity_uid(activity_type: str, description: str) -> str:
    """Generate deterministic UID using SHA256 hash of type and description."""
    content = f"{activity_type.lower().strip()}|{description.strip()}"
    return hashlib.sha256(content.encode('utf-8')).hexdigest()


def map_audience_scope(audience_target: str) -> str:
    """Map audienceTarget column to audience_scope enum."""
    if not audience_target:
        return 'all'
    
    # Normalize to lowercase for matching
    normalized = audience_target.lower().strip()
    
    mapping = {
        # Exact values from spreadsheet
        'audienceall': 'all',
        'audiencecoupleonly': 'couples',
        'audiencegrouponly': 'groups',
        
        # Alternative formats (just in case)
        'audience_all': 'all',
        'audience_couple_only': 'couples',
        'audience_group_only': 'groups',
        'couple': 'couples',
        'couples': 'couples',
        'group': 'groups',
        'groups': 'groups',
        'all': 'all',
        'both': 'all'
    }
    
    result = mapping.get(normalized, 'all')
    
    # Log warning if we're using default fallback
    if normalized and normalized not in mapping:
        print(f"⚠️  Unknown audienceTarget value: '{audience_target}', defaulting to 'all'")
    
    return result


def parse_boundaries(row: Dict[str, Any]) -> List[str]:
    """Extract hard boundary keys from columns J-Q (8 boundary columns)."""
    # These are the exact column names from the spreadsheet
    boundary_columns = [
        'hardBoundaryImpact',
        'hardBoundaryRestrain',
        'hardBoundaryBreath',
        'hardBoundaryDegrade',
        'hardBoundaryPublic',
        'hardBoundaryRecord',
        'hardBoundaryAnal',
        'hardBoundaryWatersports'
    ]
    
    boundaries = []
    for boundary_key in boundary_columns:
        value = row.get(boundary_key)
        
        # Handle different value types
        if value is None:
            continue
        
        # Convert to string and normalize
        value_str = str(value).strip().lower()
        
        # Check if this boundary is flagged
        if value_str == 'hitsboundary' or value_str == 'true' or value_str == '1':
            boundaries.append(boundary_key)
        # Ignore 'null', 'false', '0', empty strings
    
    return boundaries


def parse_anatomy(value: Any) -> List[str]:
    """Parse comma-separated anatomy values from activePlayer/partnerPlayer Must Have columns."""
    # Handle None, empty, or "null" string
    if value is None:
        return []
    
    value_str = str(value).strip()
    
    # Handle various empty/null representations
    if not value_str or value_str.lower() in ['null', 'none', 'n/a', '']:
        return []
    
    # Split by comma and clean up
    parts = [p.strip().lower() for p in value_str.split(',')]
    
    # Filter to only valid anatomy, remove duplicates
    valid = []
    seen = set()
    for part in parts:
        if part in ALLOWED_BODYPARTS and part not in seen:
            valid.append(part)
            seen.add(part)
    
    return valid


def load_enriched_data(enriched_json_path: str) -> Dict[str, Dict]:
    """Load AI-enriched activity metadata from JSON file, indexed by activity_uid."""
    if not os.path.exists(enriched_json_path):
        print(f"⚠️  No enriched data found at {enriched_json_path}")
        return {}
    
    with open(enriched_json_path, 'r') as f:
        enriched_by_row = json.load(f)
    
    # Re-index by generating activity_uid from description
    enriched_by_uid = {}
    for row_data in enriched_by_row.values():
        if 'description' in row_data and 'type' in row_data:
            uid = generate_activity_uid(row_data['type'], row_data['description'])
            enriched_by_uid[uid] = row_data
    
    print(f"✓ Loaded enriched data for {len(enriched_by_uid)} activities")
    return enriched_by_uid


def read_xlsx(filepath: str, sheet_name: str) -> List[Dict]:
    """Read XLSX file and return list of row dicts."""
    try:
        import openpyxl
    except ImportError:
        print("❌ openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        print(f"❌ Sheet '{sheet_name}' not found in workbook")
        print(f"Available sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    sheet = wb[sheet_name]
    
    # Get header row (first row)
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Read data rows
    rows = []
    for row_cells in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row_cells):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = value
        
        # Skip empty rows
        if any(row_dict.values()):
            rows.append(row_dict)
    
    wb.close()
    return rows


def read_csv(filepath: str) -> List[Dict]:
    """Read CSV file and return list of row dicts."""
    rows = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    return rows


def parse_activity_row(row: Dict, enriched_data: Optional[Dict] = None, row_num: int = 0) -> Optional[Dict[str, Any]]:
    """Parse a single row from the spreadsheet into activity data."""
    # Expected columns:
    # A: Activity Type
    # B: Activity Description
    # C: Intimacy Level
    # D: Intimacy Rating
    # E: Audience Tag (legacy - ignore)
    # F: audienceTarget (source of truth)
    # G: Audited?
    # H: activePlayer Must Have
    # I: partnerPlayer Must Have
    # J-Q: 8 boundary columns
    
    # Extract and validate core fields
    activity_type_raw = row.get('Activity Type')
    if not activity_type_raw:
        return None
    activity_type = str(activity_type_raw).strip().lower()
    
    description_raw = row.get('Activity Description')
    if not description_raw:
        return None
    description = str(description_raw).strip()
    
    intimacy_level_raw = row.get('Intimacy Level')
    intimacy_level = str(intimacy_level_raw).strip().upper() if intimacy_level_raw else ''
    
    # Get audienceTarget (column F) - this is the source of truth
    audience_target_raw = row.get('audienceTarget')
    if not audience_target_raw:
        # Fallback to legacy Audience Tag only if audienceTarget is truly empty
        audience_target_raw = row.get('Audience Tag')
    audience_target = str(audience_target_raw).strip() if audience_target_raw else ''
    
    # Validate activity type
    if activity_type not in ['truth', 'dare']:
        return None
    
    # Validate description (must be non-empty)
    if not description or len(description) < 3:
        return None
    
    # Map intimacy level to intensity and rating
    if intimacy_level not in INTIMACY_LEVEL_MAP:
        # Try to use Intimacy Rating column as fallback
        rating_col = str(row.get('Intimacy Rating', 'G')).strip().upper()
        if rating_col in ['G', 'R', 'X']:
            rating = rating_col
            # Guess intensity based on level number if present
            level_num = ''.join(filter(str.isdigit, intimacy_level))
            if level_num:
                intensity = min(int(level_num) // 2 + 1, 5)
            else:
                intensity = 2
        else:
            return None
    else:
        mapping = INTIMACY_LEVEL_MAP[intimacy_level]
        intensity = mapping['intensity']
        rating = mapping['rating']
    
    # Map audience scope
    audience_scope = map_audience_scope(audience_target)
    
    # Parse boundaries
    hard_boundaries = parse_boundaries(row)
    
    # Parse anatomy requirements
    active_anatomy = parse_anatomy(row.get('activePlayer Must Have'))
    partner_anatomy = parse_anatomy(row.get('partnerPlayer Must Have'))
    required_bodyparts = {
        "active": active_anatomy,
        "partner": partner_anatomy
    }
    
    # Build tags list
    tags = []
    if audience_target:
        tags.append(audience_target.lower())
    
    # Build script format
    script = {
        "steps": [
            {
                "actor": "A",
                "do": description
            }
        ]
    }
    
    # Generate activity UID
    activity_uid = generate_activity_uid(activity_type, description)
    
    # Base activity data
    activity_data = {
        'type': activity_type,
        'rating': rating,
        'intensity': intensity,
        'script': script,
        'tags': tags,
        'source': 'bank',
        'approved': True,
        'hard_limit_keys': [],  # Legacy field
        'audience_scope': audience_scope,
        'hard_boundaries': hard_boundaries,
        'required_bodyparts': required_bodyparts,
        'activity_uid': activity_uid,
        'source_version': 'Consolidated_ToD_Activities_v20',
        'is_active': True,
    }
    
    # Add AI-enriched metadata if available
    if enriched_data and activity_uid in enriched_data:
        enrichment = enriched_data[activity_uid]
        activity_data['power_role'] = enrichment.get('power_role')
        activity_data['preference_keys'] = enrichment.get('preference_keys', [])
        activity_data['domains'] = enrichment.get('domains', [])
        activity_data['intensity_modifiers'] = enrichment.get('intensity_modifiers', [])
        activity_data['requires_consent_negotiation'] = enrichment.get('requires_consent_negotiation', False)
    
    return activity_data


def import_activities_from_file(
    filepath: str,
    sheet_name: Optional[str] = None,
    clear_before: bool = False,
    enriched_json_path: Optional[str] = None,
    dry_run: bool = True
) -> bool:
    """
    Import activities from XLSX or CSV file into database.
    
    Args:
        filepath: Path to XLSX or CSV file
        sheet_name: Sheet name for XLSX (required if XLSX)
        clear_before: Archive all existing activities first
        enriched_json_path: Path to enriched_activities.json (optional)
        dry_run: If True, only preview without writing to database
    
    Returns:
        True if successful, False otherwise
    """
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        return False
    
    # Load enriched data if provided
    enriched_data = {}
    if enriched_json_path:
        enriched_data = load_enriched_data(enriched_json_path)
    else:
        # Auto-detect enriched_activities.json in same directory
        auto_path = Path(__file__).parent / 'enriched_activities.json'
        if auto_path.exists():
            enriched_data = load_enriched_data(str(auto_path))
    
    # Read file based on extension
    file_ext = Path(filepath).suffix.lower()
    
    if file_ext in ['.xlsx', '.xls']:
        if not sheet_name:
            print("❌ --sheet required for XLSX files")
            return False
        print(f"Reading XLSX file: {filepath}, sheet: {sheet_name}")
        rows = read_xlsx(filepath, sheet_name)
    elif file_ext == '.csv':
        print(f"Reading CSV file: {filepath}")
        rows = read_csv(filepath)
    else:
        print(f"❌ Unsupported file type: {file_ext}")
        return False
    
    print(f"✓ Loaded {len(rows)} rows from file")
    
    if dry_run:
        print("\n[DRY RUN] Preview mode - no database changes will be made")
        print("=" * 70)
    
    with app.app_context():
        # Optionally archive existing activities
        if clear_before:
            if dry_run:
                existing_count = Activity.query.filter_by(is_active=True).count()
                print(f"[DRY RUN] Would archive {existing_count} existing activities")
            else:
                print("Archiving existing active activities...")
                archived = Activity.query.filter_by(is_active=True).update({
                    'is_active': False,
                    'archived_at': datetime.utcnow()
                })
                db.session.commit()
                print(f"✓ Archived {archived} existing activities")
        
        # Track statistics
        added_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0
        
        audience_counts = {'couples': 0, 'groups': 0, 'all': 0}
        anatomy_counts = {'active': 0, 'partner': 0, 'both': 0, 'neither': 0}
        boundary_counts = {}
        
        # Get all existing activity UIDs for upsert logic
        existing_uids = set()
        if not dry_run:
            existing_activities = Activity.query.all()
            existing_uids = {a.activity_uid for a in existing_activities if a.activity_uid}
        
        # Track UIDs in import file
        import_uids = set()
        
        # Process rows in batches
        batch_size = 50
        for i in range(0, len(rows), batch_size):
            batch = rows[i:i+batch_size]
            
            for row_num, row in enumerate(batch, start=i+1):
                activity_data = parse_activity_row(row, enriched_data, row_num)
                
                if not activity_data:
                    # Only log in verbose mode or if there's actual content
                    if row.get('Activity Description'):
                        desc_preview = str(row.get('Activity Description', ''))[:40]
                        print(f"⚠️  Row {row_num}: Could not parse - {desc_preview}...")
                    skipped_count += 1
                    continue
                
                activity_uid = activity_data['activity_uid']
                import_uids.add(activity_uid)
                
                # Track statistics
                audience_counts[activity_data['audience_scope']] += 1
                
                active_parts = activity_data['required_bodyparts']['active']
                partner_parts = activity_data['required_bodyparts']['partner']
                if active_parts and partner_parts:
                    anatomy_counts['both'] += 1
                elif active_parts:
                    anatomy_counts['active'] += 1
                elif partner_parts:
                    anatomy_counts['partner'] += 1
                else:
                    anatomy_counts['neither'] += 1
                
                for boundary in activity_data['hard_boundaries']:
                    boundary_counts[boundary] = boundary_counts.get(boundary, 0) + 1
                
                if dry_run:
                    # Just count, don't insert
                    if activity_uid in existing_uids:
                        updated_count += 1
                    else:
                        added_count += 1
                    
                    # Show sample of first few activities with detailed extraction
                    if added_count + updated_count <= 5:
                        print(f"\nRow {row_num}: {activity_data['type'].upper()} - {activity_data['script']['steps'][0]['do'][:60]}...")
                        print(f"  UID: {activity_uid[:16]}...")
                        print(f"  Raw audienceTarget: '{row.get('audienceTarget', 'MISSING')}' → Mapped: '{activity_data['audience_scope']}'")
                        print(f"  Rating: {activity_data['rating']}, Intensity: {activity_data['intensity']}")
                        print(f"  Boundaries: {activity_data['hard_boundaries']}")
                        print(f"  Anatomy: active={active_parts}, partner={partner_parts}")
                        if activity_data.get('power_role'):
                            print(f"  AI Tags: power={activity_data['power_role']}, prefs={activity_data['preference_keys'][:3] if activity_data['preference_keys'] else []}")
                
                else:
                    # Upsert: check if exists by UID
                    existing = Activity.query.filter_by(activity_uid=activity_uid).first()
                    
                    try:
                        if existing:
                            # Update existing activity
                            for key, value in activity_data.items():
                                if key != 'activity_id':  # Don't update primary key
                                    setattr(existing, key, value)
                            existing.updated_at = datetime.utcnow()
                            updated_count += 1
                        else:
                            # Create new activity
                            activity = Activity(**activity_data)
                            db.session.add(activity)
                            added_count += 1
                    
                    except Exception as e:
                        print(f"❌ Row {row_num}: Error - {str(e)}")
                        error_count += 1
                        db.session.rollback()
                        continue
            
            # Commit batch
            if not dry_run:
                try:
                    db.session.commit()
                    if (i + batch_size) % 100 == 0:
                        print(f"  Processed {min(i + batch_size, len(rows))}/{len(rows)} rows...")
                except Exception as e:
                    print(f"❌ Batch commit failed: {e}")
                    db.session.rollback()
                    return False
        
        # Archive activities that were in DB but not in import file
        archived_count = 0
        if not dry_run and not clear_before:
            missing_uids = existing_uids - import_uids
            if missing_uids:
                archived_count = Activity.query.filter(
                    Activity.activity_uid.in_(missing_uids),
                    Activity.is_active == True
                ).update({
                    'is_active': False,
                    'archived_at': datetime.utcnow()
                }, synchronize_session='fetch')
                db.session.commit()
        
        # Print summary
        print("\n" + "=" * 70)
        print(f"{'[DRY RUN] ' if dry_run else ''}Import Summary")
        print("=" * 70)
        print(f"Total rows processed: {len(rows)}")
        print(f"Added: {added_count}")
        print(f"Updated: {updated_count}")
        if archived_count > 0:
            print(f"Archived: {archived_count}")
        print(f"Skipped: {skipped_count}")
        print(f"Errors: {error_count}")
        
        print(f"\nAudience Scope Distribution:")
        for scope, count in audience_counts.items():
            print(f"  {scope}: {count}")
        
        print(f"\nAnatomy Requirements:")
        for category, count in anatomy_counts.items():
            print(f"  {category}: {count}")
        
        if boundary_counts:
            print(f"\nBoundary Flags (activities with boundaries):")
            for boundary, count in sorted(boundary_counts.items(), key=lambda x: x[1], reverse=True):
                print(f"  {boundary}: {count}")
        
        if enriched_data:
            enriched_count = sum(1 for data in [parse_activity_row(r, enriched_data, 0) for r in rows] 
                               if data and data.get('power_role'))
            print(f"\nAI Enrichment: {enriched_count}/{len(rows)} activities have AI tags")
        else:
            print(f"\nAI Enrichment: Not loaded (no enriched_activities.json found)")
        
        print("=" * 70)
        
        if dry_run:
            print("\n💡 Run with --apply to execute the import")
        
        return True


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Import activities from XLSX/CSV with enhanced metadata',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Preview import (dry-run is default)
  python import_activities.py --xlsx "./Consolidated_ToD_Activities (20).xlsx" --sheet "Consolidated Activities" --dry-run
  
  # Execute import
  python import_activities.py --xlsx "./Consolidated_ToD_Activities (20).xlsx" --sheet "Consolidated Activities" --apply
  
  # Import with enrichment data
  python import_activities.py --xlsx file.xlsx --sheet "Sheet1" --enriched enriched_activities.json --apply
  
  # Archive existing activities before import
  python import_activities.py --csv activities.csv --clear-before --apply
        """
    )
    
    parser.add_argument('--xlsx', help='Path to XLSX file')
    parser.add_argument('--csv', help='Path to CSV file')
    parser.add_argument('--sheet', help='Sheet name for XLSX (required if using --xlsx)')
    parser.add_argument('--enriched', '-e', help='Path to enriched_activities.json (auto-detects if not specified)')
    parser.add_argument('--clear-before', action='store_true', help='Archive all existing activities before import')
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--dry-run', action='store_true', help='Preview import without writing to database (default)')
    group.add_argument('--apply', action='store_true', help='Execute import and write to database')
    
    args = parser.parse_args()
    
    # Validate inputs
    if args.xlsx and args.csv:
        print("❌ Error: Specify either --xlsx or --csv, not both")
        sys.exit(1)
    
    if not args.xlsx and not args.csv:
        print("❌ Error: Must specify either --xlsx or --csv")
        sys.exit(1)
    
    filepath = args.xlsx or args.csv
    
    success = import_activities_from_file(
        filepath=filepath,
        sheet_name=args.sheet,
        clear_before=args.clear_before,
        enriched_json_path=args.enriched,
        dry_run=args.dry_run
    )
    
    sys.exit(0 if success else 1)
